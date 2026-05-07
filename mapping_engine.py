import os
import json
import openpyxl
import google.generativeai as genai
from dotenv import load_dotenv

load_dotenv()

genai.configure(api_key=os.getenv("GEMINI_API_KEY"), transport="rest")
MODEL_NAME = "gemini-2.5-flash-lite"



class MappingEngine:
    def __init__(self):
        self.model = genai.GenerativeModel(MODEL_NAME)
        self.template_path = os.path.join(os.path.dirname(__file__), "Template", "Ouput template.xlsx")


    def process_task(self, task, full_workbook_data):
        """Processes a specific task from the planner's schedule."""
        mapped_data = []
        sheet_name = task.get("primary_sheet")
        target_table_names = task.get("primary_tables", [])
        context_sheets = task.get("context_sheets", [])
        instructions = task.get("processing_instructions", "")

        sheet_obj = full_workbook_data.get(sheet_name, {})
        if isinstance(sheet_obj, list):
            tables = sheet_obj
            sheet_metadata = {}
        elif isinstance(sheet_obj, dict):
            tables = sheet_obj.get('tables', [])
            sheet_metadata = sheet_obj.get('metadata', {})
        else:
            tables = []
            sheet_metadata = {}

        # Prepare context data from other sheets if needed
        extra_context = ""
        for c_sheet in context_sheets:
            c_obj = full_workbook_data.get(c_sheet, {})
            if isinstance(c_obj, list):
                c_tables = c_obj
                c_meta = {}
            elif isinstance(c_obj, dict):
                c_tables = c_obj.get('tables', [])
                c_meta = c_obj.get('metadata', {})
            else:
                c_tables = []
                c_meta = {}
                
            extra_context += f"\n--- SUPPORTING CONTEXT FROM SHEET: {c_sheet} ---\n"
            extra_context += f"Metadata: {json.dumps(c_meta)}\n"
            for t in c_tables:
                extra_context += f"Table: {t.get('name')}\n"
                # Add only first 20 rows of supporting tables to keep context manageable
                for row in t.get('data', [])[:20]:
                    extra_context += " | ".join(str(c) for c in row) + "\n"

        for table in tables:
            # Only process tables mentioned in the task, or all if none specified
            if target_table_names and table.get('name') not in target_table_names:
                continue
                
            if not target_table_names and not self._is_relevant_table(table):
                continue
                
            mapped_rows = self._map_table_with_ai(sheet_name, table, sheet_metadata, extra_context, instructions)
            
            # Apply sheet-level fallbacks if AI missed them at the table level
            for row in mapped_rows:
                if not row.get("START DATE") and sheet_metadata.get("start_date"):
                    row["START DATE"] = sheet_metadata["start_date"]
                if not row.get("EXPIRATION DATE") and sheet_metadata.get("expiration_date"):
                    row["EXPIRATION DATE"] = sheet_metadata["expiration_date"]
                if not row.get("PROVIDER") and sheet_metadata.get("provider"):
                    row["PROVIDER"] = sheet_metadata["provider"]
                if not row.get("COMMODITY") and sheet_metadata.get("commodity"):
                    row["COMMODITY"] = sheet_metadata["commodity"]
                if not row.get("SERVICE NAME") and sheet_metadata.get("service_name"):
                    row["SERVICE NAME"] = sheet_metadata["service_name"]
                
                # Dynamic fallback for Origin and Destination from global metadata
                if not row.get("ORIGIN LOCATION") and sheet_metadata.get("global_origin"):
                    row["ORIGIN LOCATION"] = sheet_metadata["global_origin"]
                if not row.get("ORIGIN") and sheet_metadata.get("global_origin"):
                    row["ORIGIN"] = sheet_metadata["global_origin"]
                if not row.get("DESTINATION LOCATION") and sheet_metadata.get("global_destination"):
                    row["DESTINATION LOCATION"] = sheet_metadata["global_destination"]
                if not row.get("DESTINATION") and sheet_metadata.get("global_destination"):
                    row["DESTINATION"] = sheet_metadata["global_destination"]

                # Python fallback to guarantee valid Ports are never missing by using Location names directly
                if not row.get("ORIGIN PORT") and row.get("ORIGIN LOCATION"):
                    row["ORIGIN PORT"] = str(row["ORIGIN LOCATION"]).strip().upper()
                if not row.get("DESTINATION PORT") and row.get("DESTINATION LOCATION"):
                    row["DESTINATION PORT"] = str(row["DESTINATION LOCATION"]).strip().upper()

                # Enforce that the sheet name is strictly written in the NOTES / NOTAS field
                if not row.get("NOTES"):
                    row["NOTES"] = str(sheet_name).strip()
                else:
                    existing_notes = str(row.get("NOTES")).strip()
                    if str(sheet_name).strip() not in existing_notes:
                        row["NOTES"] = f"[{str(sheet_name).strip()}] {existing_notes}"
                
            mapped_data.extend(mapped_rows)

        # Separate LLM call to find additional charges and surcharges (isolated surcharge discovery path)
        try:
            print(f"Running separate surcharge extraction LLM call for sheet: {sheet_name}...")
            additional_rows = self._extract_additional_surcharges(sheet_name, tables, sheet_metadata, extra_context)
            for row in additional_rows:
                if not row.get("START DATE") and sheet_metadata.get("start_date"):
                    row["START DATE"] = sheet_metadata["start_date"]
                if not row.get("EXPIRATION DATE") and sheet_metadata.get("expiration_date"):
                    row["EXPIRATION DATE"] = sheet_metadata["expiration_date"]
                if not row.get("PROVIDER") and sheet_metadata.get("provider"):
                    row["PROVIDER"] = sheet_metadata["provider"]
                if not row.get("COMMODITY") and sheet_metadata.get("commodity"):
                    row["COMMODITY"] = sheet_metadata["commodity"]
                if not row.get("SERVICE NAME") and sheet_metadata.get("service_name"):
                    row["SERVICE NAME"] = sheet_metadata["service_name"]
                
                if not row.get("ORIGIN LOCATION") and sheet_metadata.get("global_origin"):
                    row["ORIGIN LOCATION"] = sheet_metadata["global_origin"]
                if not row.get("ORIGIN") and sheet_metadata.get("global_origin"):
                    row["ORIGIN"] = sheet_metadata["global_origin"]
                if not row.get("DESTINATION LOCATION") and sheet_metadata.get("global_destination"):
                    row["DESTINATION LOCATION"] = sheet_metadata["global_destination"]
                if not row.get("DESTINATION") and sheet_metadata.get("global_destination"):
                    row["DESTINATION"] = sheet_metadata["global_destination"]

                if not row.get("ORIGIN PORT") and row.get("ORIGIN LOCATION"):
                    row["ORIGIN PORT"] = str(row["ORIGIN LOCATION"]).strip().upper()
                if not row.get("DESTINATION PORT") and row.get("DESTINATION LOCATION"):
                    row["DESTINATION PORT"] = str(row["DESTINATION LOCATION"]).strip().upper()

                if not row.get("NOTES"):
                    row["NOTES"] = str(sheet_name).strip()
                else:
                    existing_notes = str(row.get("NOTES")).strip()
                    if str(sheet_name).strip() not in existing_notes:
                        row["NOTES"] = f"[{str(sheet_name).strip()}] {existing_notes}"
            
            print(f"Surcharge extraction complete. Appending {len(additional_rows)} surcharges to results.")
            mapped_data.extend(additional_rows)
        except Exception as e:
            print(f"Error merging additional surcharges for sheet {sheet_name}: {e}")
            
        return mapped_data

    def _extract_additional_surcharges(self, sheet_name, tables, sheet_metadata=None, extra_context=""):
        """Specialized, separate LLM call to extract any additional charges and surcharges from the sheet tables/metadata."""
        # Convert all tables to a concise summary of fields/rows to look for surcharges
        formatted_tables_data = ""
        for table in tables:
            formatted_tables_data += f"\nTable: {table.get('name', 'Unknown')}\n"
            headers = table.get('headers', [])
            formatted_tables_data += "Headers: " + " | ".join(str(h) for h in headers if h is not None) + "\n"
            for row in table.get('data', [])[:30]:  # Give enough rows to inspect surcharges
                formatted_tables_data += " | ".join(str(cell) if cell is not None else "" for cell in row) + "\n"

        prompt = f"""
        You are a highly experienced Logistics Surcharge Auditor. Your task is to extract ANY and ALL additional charges, surcharges, local fees, demurrage/detention fees, seal fees, VGM fees, or administrative fees mentioned in this sheet.
        
        Sheet Name: {sheet_name}
        
        --- GLOBAL SHEET METADATA ---
        {json.dumps(sheet_metadata if sheet_metadata else {})}
        
        --- CROSS-SHEET SUPPORTING CONTEXT ---
        {extra_context}
        
        --- ALL SHEET TABLES DATA ---
        {formatted_tables_data}
        
        --- EXTRACTION INSTRUCTIONS ---
        1. Identify any surcharge rows, columns, or tables (e.g. BAF, CAF, THC, ISPS, LSS, IMO, Demurrage, Detentions, Seal/Precinto, VGM, T-3, etc.).
        2. Generate a separate JSON object (row) for EACH individual charge.
        3. Do NOT include standard ocean freight rates here (only extract additional charges and surcharges).
        4. Standardize the "CHARGE" column using standard logistics codes: e.g. THC, BAF, LSS, IMO, EUIS, ERC, OWS, PSS, SCMC, VGM, SEAL FEE, DEMURRAGE, DETENTION, CANCELACION.
        5. Strictly populate "START DATE" and "EXPIRATION DATE" in YYYY-MM-DD format using global metadata or table context.
        6. Strictly populate "PROVIDER" with the carrier/provider name.
        7. Format all pricing columns (20DRY, 40DRY, etc.) and Transit Time as INTEGERS.
        8. CRITICAL LIST FORMATTING RULE: For any column containing multiple items or lists (such as INCLUDED CHARGES, REMARKS, or NOTES), always separate the values with a pipe "|" instead of a comma ",". For example, use "THC | SEC | BAF" instead of "THC, SEC, BAF".

        
        --- OUTPUT FORMAT ---
        Return ONLY a JSON array of objects. Each object MUST match these keys:
        "ORIGIN LOCATION", "ORIGIN PORT", "DESTINATION PORT", "DESTINATION LOCATION", "ORIGIN", "DESTINATION",
        "CHARGE TYPE", "CHARGE", "RATE BASIS", "CURRENCY", 
        "20DRY", "40DRY", "40HDRY", "45HDRY", "40NOR", "20RF", "40HCRF", "45RF", "20OT", "40OT", "40HCOT", "20FR", "40FR", "40HCFR", "20TK", "40TK",
        "PAYMENT TERM", "PROVIDER", "LIMITS", "START DATE", "EXPIRATION DATE", "VIA", "TRANSIT TIME", "COMMODITY", "SERVICE NAME", 
        "INCLUDED CHARGES", "REMARKS", "MODE OF TRANSPORT", "EXCEPTIONS ORIGIN", "EXCEPTIONS DESTINATION", "NOTES", "RATE OVER"
        
        CRITICAL: Return ONLY the JSON array. Do not include conversational text or Markdown formatting.
        """
        
        try:
            response = self.model.generate_content(prompt)
            text = response.text.strip()
            
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0]
            elif "```" in text:
                text = text.split("```")[1].split("```")[0]
            
            start_idx = text.find('[')
            end_idx = text.rfind(']') + 1
            if start_idx != -1 and end_idx != 0:
                json_str = text[start_idx:end_idx].strip()
                json_str = json_str.replace(",\n]", "\n]").replace(",]", "]")
                return json.loads(json_str)
            return []
        except Exception as e:
            print(f"Surcharge extraction AI failed for sheet {sheet_name}: {e}")
            return []

    def _is_relevant_table(self, table):
        # Basic heuristic: does it have price-like headers or data?
        headers = table.get('headers', [])
        headers_str = " ".join(str(h) for h in headers if h is not None).lower()
        relevant_keywords = [
            '20', '40', 'rate', 'price', 'charge', 'cost', 'dry', 'rf', 'ocean', 'freight',
            'flete', 'recargo', 'importe', 'precio', 'gasto', 'tasa', 'tasas', 'usd', 'eur', 'gbp', '$', '€'
        ]
        return any(kw in headers_str for kw in relevant_keywords)

    def _map_table_with_ai(self, sheet_name, table, sheet_metadata=None, extra_context="", instructions=""):
        # Format table data as pipe-separated rows for better AI spatial reasoning
        data_rows = table.get('data', [])
        context_rows = table.get('context', [])
        
        formatted_data = ""
        for row in data_rows[:80]:
            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
            while clean_row and clean_row[-1] == "":
                clean_row.pop()
            if any(cell != "" for cell in clean_row):
                formatted_data += " | ".join(clean_row) + "\n"

        formatted_context = ""
        for row in context_rows:
            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
            while clean_row and clean_row[-1] == "":
                clean_row.pop()
            if any(cell != "" for cell in clean_row):
                formatted_context += " | ".join(clean_row) + "\n"


        prompt = f"""
        You are a highly experienced Logistics Data Agent. Your task is to extract rate data from an Excel table and map it to our standard output format.
        
        Sheet Name: {sheet_name}
        Table Name: {table.get('name', 'Unknown')}
        Table Range: {table.get('range', 'Unknown')}
        Table Headers: {" | ".join(str(h) for h in table.get('headers', []) if h is not None)}
        
        --- GLOBAL SHEET METADATA (Fallback) ---
        {json.dumps(sheet_metadata if sheet_metadata else {})}

        --- STRATEGIC PROCESSING INSTRUCTIONS ---
        {instructions}

        --- CROSS-SHEET SUPPORTING CONTEXT ---
        {extra_context}

        --- CONTEXT (Rows above the table - USE FOR ORIGIN/DESTINATION/DATES) ---
        {formatted_context}

        --- TABLE DATA SNIPPET (Pipe Separated) ---
        {formatted_data}

        --- GOLDEN RULES ---
        1. ORIGIN/DESTINATION (MANDATORY ON EVERY ROW): 
           - EVERY SINGLE ROW in the output JSON array MUST have a valid "ORIGIN PORT" and "DESTINATION PORT" (for FCL) or "ORIGIN" and "DESTINATION" (for ARB). Leaving these fields empty or null is an absolute system failure.
           - If not explicitly listed inside the row, look at the table title, surrounding 'CONTEXT' rows, 'Table Name', 'Sheet Name', or supporting context to find the origin and destination, and write them for EVERY SINGLE LINE.
           - Extract the EXACT port name or city name from the contract (e.g. VALENCIA, BRISBANE, SHANGHAI) and write it directly under "ORIGIN PORT" / "DESTINATION PORT". DO NOT translate these into 5-letter UN/LOCODEs unless the contract itself lists them as LOCODEs.
        2. IMPORT/EXPORT LOGIC: If Origin is Spain, Portugal, or Italy, it is EXPORT. If Destination is Spain, Portugal, or Italy, it is IMPORT.
        3. CONTAINER TYPES & MULTIPLE CHARGES (ROW SPLITTING): 
           - Map prices to the standard 16 columns (20DRY, 40DRY, etc.).
           - Values MUST be INTEGERS. Remove all currency symbols, commas, or text like "Consultar". If not a number, leave null.
           - CRITICAL ROW SPLITTING RULE: If a table lists multiple distinct charges as separate columns (e.g. separate columns for Freight, BAF, THC, LSS, ERC, etc.) or if there are separate surcharge tables, you MUST generate a SEPARATE JSON object (row) for EACH individual charge. 
             For example, instead of one combined row, generate one row for "FREIGHT" with its prices, another row for "BAF" with its prices, and another row for "THC" with its prices.
             Never combine multiple active charges into a single row unless a charge is explicitly noted as "INCLUDED" in the rate.
        4. CHARGE MAPPING (STANDARD CODES):
           - Standardize the "CHARGE" name to logical logistics codes (e.g., FREIGHT, THC, THCO, BAF, LSS, IMO, EUIS, ERC, OWS, PSS, SCMC, EBS, ENS, VGM, SEAL FEE, WRC).
           - Do NOT copy verbatim raw text like "Emisión certificado VGM" or "Precinto (seal fee)"; standardize them to standard codes like "VGM" or "SEAL FEE".
        5. CHARGE TYPE & SHEET SELECTION (MODE OF TRANSPORT): 
           - Identify if the charge is "Ocean Freight", "Origin Charge", "Destination Charge", or "Surcharge" based on context.
           - Crucial: Correctly identify the "MODE OF TRANSPORT" (e.g., "SEA", "OCEAN", "FCL", "ROAD", "RAIL", "TRUCK", "TRAIN", "ARBITRARY"). 
             If the rate is land-based (Road/Rail arbitraries or inland feedering), strictly set MODE OF TRANSPORT to "ROAD" or "RAIL" so our system routes it to the "Arbitraries" sheet. Otherwise, set it to "SEA" or "OCEAN" to route it to the "Fletes y Recargos" sheet.
        6. PROVIDER (CRITICAL): 
           - You MUST strictly find the actual provider/carrier (e.g., Evergreen, Cosco, MSC, CMA CGM, Hapag-Lloyd) from the contract sheet context, logos, headers, or metadata, and fill it in the "PROVIDER" column.
           - DO NOT leave this empty or default it. Every row must strictly contain the actual provider.
        7. RATE BASIS: Map to standard: PER_CONTAINER, PER_TEU, PER_BL, or PER_SHIPMENT.
        8. CURRENCY: Extract the ISO currency code (USD, EUR, etc.) from the table or context.
        9. DATES (START DATE & EXPIRATION DATE - MANDATORY): 
           - EVERY SINGLE ROW in the output JSON array MUST have a valid "START DATE" and "EXPIRATION DATE" in "YYYY-MM-DD" format. Leaving these fields empty or null is an absolute system failure.
           - Be a timeline detective: look at the top rows, footers, sheet names, table headers, or 'GLOBAL SHEET METADATA' to find when these rates are valid.
           - If a range is mentioned (e.g., "Effective April 1st, 2026", "Valid through Dec 31, 2026"), convert and write them as "YYYY-MM-DD" for EVERY SINGLE LINE without exception.
        10. TRANSIT TIME: MUST be a single INTEGER. If a range is given, pick the representative number.
        11. COMMODITY: Extract specific commodity restrictions (e.g., FAK, Food Grade).
        12. INCLUDED CHARGES: Identify any surcharges listed as "Included" in the rate or notes. Use pipe-delimited codes.
        13. NOTES: You MUST strictly write the original sheet name '{sheet_name}' in this field for every single row.
        14. NULL FIELDS: If a field is not found in the file, return null. DO NOT invent data.
        15. CRITICAL LIST FORMATTING RULE: For any column containing multiple items or lists (such as INCLUDED CHARGES, REMARKS, or NOTES), always separate the values with a pipe "|" instead of a comma ",". For example, use "THC | SEC | BAF" instead of "THC, SEC, BAF".

        --- OUTPUT FORMAT ---
        Return ONLY a JSON array of objects. Each object MUST match these keys:
        "ORIGIN LOCATION", "ORIGIN PORT", "DESTINATION PORT", "DESTINATION LOCATION", "ORIGIN", "DESTINATION",
        "CHARGE TYPE", "CHARGE", "RATE BASIS", "CURRENCY", 
        "20DRY", "40DRY", "40HDRY", "45HDRY", "40NOR", "20RF", "40HCRF", "45RF", "20OT", "40OT", "40HCOT", "20FR", "40FR", "40HCFR", "20TK", "40TK",
        "PAYMENT TERM", "PROVIDER", "LIMITS", "START DATE", "EXPIRATION DATE", "VIA", "TRANSIT TIME", "COMMODITY", "SERVICE NAME", 
        "INCLUDED CHARGES", "REMARKS", "MODE OF TRANSPORT", "EXCEPTIONS ORIGIN", "EXCEPTIONS DESTINATION", "NOTES", "RATE OVER"

        CRITICAL: 
        - Return ONLY the JSON. No conversational text.
        - NUMERIC DATA: Ensure all price columns and Transit Time are INTEGERS.
        - DATES: Ensure they are in YYYY-MM-DD format.
        - START DATE & EXPIRATION DATE ARE MANDATORY ON EVERY ROW: Every single row must strictly contain a valid START DATE and EXPIRATION DATE in YYYY-MM-DD format. Look everywhere in the sheet metadata and context to extract them. DO NOT leave them empty!
        - ORIGIN/DESTINATION PORT EXTRACTION IS ABSOLUTELY MANDATORY: You must identify and write the origin and destination ports for EVERY SINGLE row. Search all metadata, sheet names, table names, context rows, and headers to find them. Do not leave them empty!
        """

        try:
            response = self.model.generate_content(prompt)
            text = response.text.strip()
            
            # Extract JSON more robustly
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0]
            elif "```" in text:
                text = text.split("```")[1].split("```")[0]
            
            start_idx = text.find('[')
            end_idx = text.rfind(']') + 1
            if start_idx != -1 and end_idx != 0:
                json_str = text[start_idx:end_idx].strip()
                # Remove common AI malformations like trailing commas before closing bracket
                json_str = json_str.replace(",\n]", "\n]").replace(",]", "]")
                return json.loads(json_str)
            
            return []
        except Exception as e:
            print(f"Mapping AI failed for table {table.get('name')}: {e}")
            return []

    def write_to_template(self, mapped_rows, output_path):
        """Writes the mapped rows into the Excel template, splitting by transport mode."""
        import traceback
        print(f"DEBUG: Starting write_to_template with {len(mapped_rows)} rows to output_path: {output_path}")
        
        # Standard fallback headers in case template cannot be loaded
        STANDARD_HEADERS = [
            "ORIGIN LOCATION", "ORIGIN PORT", "DESTINATION PORT", "DESTINATION LOCATION", "ORIGIN", "DESTINATION",
            "CHARGE TYPE", "CHARGE", "RATE BASIS", "CURRENCY", 
            "20DRY", "40DRY", "40HDRY", "45HDRY", "40NOR", "20RF", "40HCRF", "45RF", "20OT", "40OT", "40HCOT", "20FR", "40FR", "40HCFR", "20TK", "40TK",
            "PAYMENT TERM", "PROVIDER", "LIMITS", "START DATE", "EXPIRATION DATE", "VIA", "TRANSIT TIME", "COMMODITY", "SERVICE NAME", 
            "INCLUDED CHARGES", "REMARKS", "MODE OF TRANSPORT", "EXCEPTIONS ORIGIN", "EXCEPTIONS DESTINATION", "NOTES", "RATE OVER"
        ]
        
        try:
            wb = None
            if os.path.exists(self.template_path):
                try:
                    print(f"DEBUG: Loading template from: {self.template_path}")
                    wb = openpyxl.load_workbook(self.template_path)
                except Exception as load_err:
                    print(f"DEBUG: Failed to load template file via openpyxl: {load_err}")
            
            if wb is None:
                print("DEBUG: Template file not found or load failed. Generating empty workbook from code definitions...")
                wb = openpyxl.Workbook()
                # Create standard sheets
                fcl_ws = wb.active
                fcl_ws.title = "Fletes y Recargos "
                arb_ws = wb.create_sheet(title="Arbitraries")
                
                # Write headers to both sheets
                for col_idx, header in enumerate(STANDARD_HEADERS, 1):
                    fcl_ws.cell(row=1, column=col_idx, value=header)
                    arb_ws.cell(row=1, column=col_idx, value=header)
            
            # Sheet mapping
            # Note: "Fletes y Recargos " has a trailing space in the template
            sheets = {
                "FCL": "Fletes y Recargos ", 
                "ARB": "Arbitraries"
            }
            
            # Prepare column maps for both sheets
            col_maps = {}
            for key, sheet_name in sheets.items():
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    headers = [cell.value for cell in ws[1]]
                    cmap = {str(header).strip().upper(): idx + 1 for idx, header in enumerate(headers) if header}
                    col_maps[key] = (ws, cmap)
                else:
                    print(f"DEBUG: Sheet '{sheet_name}' not found in workbook. Creating it and writing standard headers...")
                    ws = wb.create_sheet(title=sheet_name)
                    for col_idx, header in enumerate(STANDARD_HEADERS, 1):
                        ws.cell(row=1, column=col_idx, value=header)
                    cmap = {str(header).strip().upper(): idx + 1 for idx, header in enumerate(STANDARD_HEADERS) if header}
                    col_maps[key] = (ws, cmap)

            print(f"DEBUG: Column mapping completed successfully. Now writing rows...")
            for row_idx, row_data in enumerate(mapped_rows, 1):
                mode = str(row_data.get("MODE OF TRANSPORT", "")).upper()
                
                # Determine target sheet
                # Arbitraries are usually Road/Rail or specifically marked
                if any(x in mode for x in ["ROAD", "RAIL", "TRUCK", "TRAIN", "ARBITRARY"]):
                    target_key = "ARB"
                else:
                    target_key = "FCL"
                
                if target_key in col_maps:
                    ws, cmap = col_maps[target_key]
                    start_row = ws.max_row + 1
                    for header, value in row_data.items():
                        clean_header = str(header).strip().upper()
                        if clean_header in cmap:
                            # Handle potential list values from AI (e.g., list of included charges)
                            if isinstance(value, list):
                                value = " | ".join(map(str, value))
                            
                            # Ensure numeric values are actually numeric for Excel
                            if clean_header in ["20DRY", "40DRY", "40HDRY", "45HDRY", "40NOR", "20RF", "40HCRF", "45RF", "20OT", "40OT", "40HCOT", "20FR", "40FR", "40HCFR", "20TK", "40TK", "TRANSIT TIME"]:
                                try:
                                    if value is not None and str(value).strip():
                                        # Remove any non-numeric characters except decimal point
                                        clean_val = "".join(c for c in str(value) if c.isdigit() or c == '.')
                                        value = float(clean_val) if '.' in clean_val else int(clean_val)
                                    else:
                                        value = None
                                except:
                                    value = None
                            
                            # Ensure dates are datetime objects
                            if clean_header in ["START DATE", "EXPIRATION DATE"] and value:
                                try:
                                    from datetime import datetime
                                    if isinstance(value, str):
                                        value = datetime.strptime(value.split('T')[0], "%Y-%m-%d")
                                except:
                                    pass

                            ws.cell(row=start_row, column=cmap[clean_header], value=value)
            
            print(f"DEBUG: Saving final mapped workbook to: {output_path}")
            wb.save(output_path)
            print("DEBUG: Workbook saved successfully!")
            return True
        except Exception as e:
            print(f"CRITICAL ERROR writing to template: {e}")
            print(traceback.format_exc())
            return False


mapping_engine = MappingEngine()
