import openpyxl

template_path = r"c:\Users\AxeGlobal Ai\Desktop\test ground\SAMPLE SHEET\Ouput template.xlsx"
wb = openpyxl.load_workbook(template_path)
ws = wb['Fletes y Recargos ']
print("Template 'Fletes y Recargos ' column headers:")
headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
for idx, h in enumerate(headers):
    print(f"Col {idx+1}: {h}")
