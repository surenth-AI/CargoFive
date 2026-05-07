import openpyxl
file_path = r"C:\Users\AxeGlobal Ai\Desktop\test ground\SAMPLE SHEET\46566_FCL_TRANSPORTES GRUPO CALICHE_contract.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
for name in ["EXP BARCELONA", "EXP CASTELLON"]:
    if name in wb.sheetnames:
        sheet = wb[name]
        print(f"\n--- {name} ---")
        rows = list(sheet.iter_rows(min_row=1, max_row=20, min_col=1, max_col=12, values_only=True))
        for i, r in enumerate(rows, 1):
            if any(x is not None for x in r):
                print(f"Row {i:02d}:", [x for x in r if x is not None])
