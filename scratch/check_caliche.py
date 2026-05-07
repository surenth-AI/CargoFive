import openpyxl
import os

file_path = r"C:\Users\AxeGlobal Ai\Desktop\test ground\SAMPLE SHEET\46566_FCL_TRANSPORTES GRUPO CALICHE_contract.xlsx"
if not os.path.exists(file_path):
    print("File does not exist:", file_path)
else:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    for name in ["EXP BARCELONA", "EXP CASTELLON", "EXP VALENCIA"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            print(f"\n--- {name} ---")
            print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")
            rows = []
            for r in list(sheet.iter_rows(min_row=1, max_row=45, min_col=1, max_col=15, values_only=True)):
                rows.append(r)
            for i, r in enumerate(rows[:45], 1):
                # Filter out all-None rows
                if any(x is not None for x in r):
                    non_none = [x for x in r if x is not None]
                    print(f"Row {i:02d}:", non_none[:12])
        else:
            print(f"Sheet {name} not found in workbook.")
