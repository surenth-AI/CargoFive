import os
import pandas as pd
import openpyxl
import google.generativeai as genai
from dotenv import load_dotenv
import json
import time
from concurrent.futures import ThreadPoolExecutor

from openpyxl.utils.cell import range_boundaries, column_index_from_string

load_dotenv()

genai.configure(api_key=os.getenv("GEMINI_API_KEY"), transport="rest")
MODEL_NAME = "gemini-2.5-flash-lite"



class DiscoveryEngine:
    def __init__(self):
        self.model = genai.GenerativeModel(MODEL_NAME)

    def process_excel(self, file_path):
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            results = {}

            def process_single_sheet(sheet_name):
                try:
                    sheet = workbook[sheet_name]
                    
                    # 1. Map merged cells for consistent data extraction
                    merged_map = self._get_merged_cells_map(sheet)
                    
                    # 2. Extract structural data (with merged cells handled)
                    sheet_data = self._extract_sheet_data(sheet, merged_map)
                    
                    if not sheet_data:
                        return sheet_name, {
                            "metadata": {},
                            "tables": []
                        }

                    # We run these two independent AI calls concurrently inside each sheet worker!
                    with ThreadPoolExecutor(max_workers=2) as inner_executor:
                        meta_future = inner_executor.submit(self._get_sheet_metadata, sheet_name, sheet_data)
                        table_future = inner_executor.submit(self._analyze_with_ai, sheet_name, sheet_data)
                        
                        sheet_metadata = meta_future.result()
                        table_metadata = table_future.result()
                    
                    # 5. Populate full data programmatically
                    tables = self._populate_table_data(table_metadata, sheet_data)
                    
                    return sheet_name, {
                        "metadata": sheet_metadata,
                        "tables": tables
                    }
                except Exception as sheet_err:
                    print(f"Error processing sheet '{sheet_name}' in parallel: {sheet_err}")
                    return sheet_name, {
                        "metadata": {},
                        "tables": []
                    }

            sheet_names = workbook.sheetnames
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(process_single_sheet, name): name for name in sheet_names}
                for future in futures:
                    name = futures[future]
                    try:
                        name, data = future.result()
                        results[name] = data
                    except Exception as e:
                        print(f"Error reading parallel discovery result for sheet {name}: {e}")
                        results[name] = {"metadata": {}, "tables": []}

            return results
        except Exception as e:
            print(f"Error processing excel: {e}")
            raise e

    def _get_sheet_metadata(self, sheet_name, data_rows):
        """Dedicated AI call to find global sheet info like Dates and Provider."""
        prompt = f"""
        Analyze the first 200 rows of this Excel sheet named '{sheet_name}'.
        Find global metadata that applies to all rates in this sheet.
        
        Look for:
        1. START DATE: When these rates become valid.
        2. EXPIRATION DATE: When these rates expire.
        3. PROVIDER: Which company issued this quote (e.g., MSC, Maersk, Kuehne+Nagel).
        4. COMMODITY: (e.g., FAK, General Cargo).
        5. SERVICE NAME: (e.g., Silk Route, MED-USA).
        6. GLOBAL ORIGIN: Any global origin port/city (e.g., Barcelona, ESMAD).
        7. GLOBAL DESTINATION: Any global destination port/city (e.g., Shanghai, CNSHA).

        Return ONLY a JSON object:
        {{
            "start_date": "YYYY-MM-DD or null",
            "expiration_date": "YYYY-MM-DD or null",
            "provider": "string or null",
            "commodity": "string or null",
            "service_name": "string or null",
            "global_origin": "string or null",
            "global_destination": "string or null"
        }}

        Sheet Data Snippet:
        {json.dumps(data_rows[:200])}
        """
        attempts = 5
        for attempt in range(attempts):
            try:
                response = self.model.generate_content(prompt)
                text = response.text
                if "```json" in text:
                    text = text.split("```json")[1].split("```")[0]
                start_idx = text.find('{')
                end_idx = text.rfind('}') + 1
                if start_idx != -1:
                    return json.loads(text[start_idx:end_idx])
                return {}
            except Exception as e:
                err_str = str(e)
                if "429" in err_str or "ResourceExhausted" in err_str or "Quota exceeded" in err_str:
                    wait_time = (2 ** attempt) * 6
                    print(f"Metadata AI hit 429 for '{sheet_name}'. Waiting {wait_time}s (attempt {attempt+1}/{attempts})...")
                    time.sleep(wait_time)
                else:
                    print(f"Sheet Metadata AI failed for '{sheet_name}': {e}")
                    return {}
        return {}

    def _get_merged_cells_map(self, sheet):
        """Creates a lookup map for merged cells. Only the top-left cell gets the value, others get a [MERGED] token."""
        merged_map = {}
        for merged_range in sheet.merged_cells.ranges:
            for r_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for c_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    if r_idx == merged_range.min_row and c_idx == merged_range.min_col:
                        continue
                    merged_map[(r_idx, c_idx)] = "[MERGED]"
        return merged_map

    def _extract_sheet_data(self, sheet, merged_map):
        data = []
        max_row = sheet.max_row
        sample_limit = 3000

        # Scan first 300 rows to find actual maximum column containing data
        actual_max_col = 1
        for row in sheet.iter_rows(min_row=1, max_row=min(max_row, 300), values_only=True):
            for c_idx, val in enumerate(row, 1):
                if val is not None and c_idx > actual_max_col:
                    actual_max_col = c_idx
        
        max_col = min(sheet.max_column, actual_max_col + 2) # Buffer of 2 columns

        for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col), 1):
            if r_idx > sample_limit: break
            
            row_data = []
            has_value = False
            for c_idx, cell in enumerate(row, 1):
                if (r_idx, c_idx) in merged_map:
                    val = merged_map[(r_idx, c_idx)]
                else:
                    val = cell.value
                
                if val is not None:
                    has_value = True
                
                if hasattr(val, 'isoformat'):
                    row_data.append(val.isoformat())
                else:
                    row_data.append(val)
            
            # Trim trailing None values
            while row_data and row_data[-1] is None:
                row_data.pop()
                
            if has_value:
                data.append(row_data)
        
        return data


    def _analyze_with_ai(self, sheet_name, data_rows):
        def json_serial(obj):
            if hasattr(obj, 'isoformat'):
                return obj.isoformat()
            return str(obj)

        prompt = f"""
        You are a world-class data structure specialist. I have an Excel sheet named '{sheet_name}'.
        
        Your task:
        1. Find ALL distinct data tables in the provided data snippet.
        2. Identify the EXACT boundaries (Excel range like A1:G2500).
        3. Identify the logical headers for each table.
        4. FIND TABLE NAME: Look for a title or name in the rows immediately ABOVE the table headers. 
           - Mostly, tables have names like "Import Customer Service" or "Sales Report 2024".
           - Use the actual name found in the sheet if present.
           - ONLY if no name is found in the sheet, generate a highly relevant one.
        5. Resolve Merged Cells: If a column is [MERGED], it belongs to the cell to its left.
        
        CRITICAL: DO NOT return the data rows. ONLY return the metadata. DO NOT include empty cells or null values in the JSON structure.

        Return the result STRICTLY as a valid JSON array of objects:
        [
            {{
                "range": "A1:G25",
                "headers": ["POD", "20'", "40'"],
                "table_name": "Actual Name from Sheet or Relevant Name",
                "type": "Table Description",
                "start_row": 1, 
                "end_row": 25,
                "start_col": "A",
                "end_col": "G"
            }}
        ]

        Sheet Data Snippet (with [MERGED] tokens):
        {json.dumps(data_rows[:300], default=json_serial)} 
        """

        attempts = 5
        for attempt in range(attempts):
            try:
                response = self.model.generate_content(prompt)
                text = response.text
                if "```json" in text:
                    text = text.split("```json")[1].split("```")[0]
                elif "```" in text:
                    text = text.split("```")[1].split("```")[0]
                
                start_idx = text.find('[')
                end_idx = text.rfind(']') + 1
                if start_idx != -1 and end_idx != 0:
                    text = text[start_idx:end_idx]
                    
                return json.loads(text.strip())
            except Exception as e:
                err_str = str(e)
                if "429" in err_str or "ResourceExhausted" in err_str or "Quota exceeded" in err_str:
                    wait_time = (2 ** attempt) * 6
                    print(f"Table Discovery AI hit 429 for '{sheet_name}'. Waiting {wait_time}s (attempt {attempt+1}/{attempts})...")
                    time.sleep(wait_time)
                else:
                    print(f"AI Analysis failed for sheet {sheet_name}: {e}")
                    return []
        return []

    def _populate_table_data(self, table_metadata, sheet_data):
        """Extracts the actual data from sheet_data using the AI-provided range metadata."""
        populated_tables = []
        
        for meta in table_metadata:
            try:
                # Use openpyxl utility to parse range
                # meta['range'] might be "A1:G2500"
                if 'range' in meta:
                    min_col, min_row, max_col, max_row = range_boundaries(meta['range'])
                else:
                    # Fallback to individual fields
                    min_row = int(meta.get('start_row', 1))
                    max_row = int(meta.get('end_row', len(sheet_data)))
                    min_col = column_index_from_string(meta.get('start_col', 'A'))
                    max_col = column_index_from_string(meta.get('end_col', 'Z'))

                # Slice sheet_data
                # Note: openpyxl uses 1-based indexing, sheet_data is 0-based
                table_rows = []
                for r in range(min_row - 1, min_row + 100): # Limit to 100 rows for PREVIEW in UI
                     # Wait, user said "Exactly like in the sheet", but showing 2500 rows in HTML will crash the browser.
                     # I will return the first 100 rows and a note that it's truncated for view.
                     pass
                
                # Actually, let's return ALL data as requested, but be careful.
                # If it's 2500 rows, we return it.
                table_rows = [row[min_col-1:max_col] for row in sheet_data[min_row-1:max_row]]
                
                # Clean up [MERGED] tokens in the final output for the user
                cleaned_rows = []
                for row in table_rows:
                    cleaned_rows.append([None if cell == "[MERGED]" else cell for cell in row])

                # Get context (10 rows above headers)
                context_rows = []
                context_start = max(0, min_row - 11)
                for row in sheet_data[context_start : min_row - 1]:
                    context_rows.append([str(cell) if cell is not None and cell != "[MERGED]" else "" for cell in row])

                populated_tables.append({
                    "range": meta.get('range', f"{meta.get('start_col')}{min_row}:{meta.get('end_col')}{max_row}"),
                    "headers": meta.get('headers', []),
                    "data": cleaned_rows,
                    "context": context_rows, # Added context for AI
                    "name": meta.get('table_name', meta.get('type', 'Data Table')),
                    "type": meta.get('type', 'Data Table')
                })
            except Exception as e:
                print(f"Error populating table data: {e}")
                continue
                
        return populated_tables

discovery_engine = DiscoveryEngine()
